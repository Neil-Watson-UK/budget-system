# create_excel_template.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
import requests
from datetime import datetime
import os

class BudgetExcelGenerator:
    def __init__(self, api_url="https://eposaudioevents.com/budgets/excel_api.php"):
        self.api_url = api_url
        self.filters = {
            'region': 'ALL',
            'status': 'ALL',
            'format': 'json'
        }
        
    def fetch_data(self):
        """Fetch data from the API"""
        try:
            response = requests.get(self.api_url, params=self.filters, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            if data.get('success'):
                df = pd.DataFrame(data['data'])
                metadata = data.get('metadata', {})
                return df, metadata
            else:
                raise Exception(f"API Error: {data.get('error', 'Unknown error')}")
                
        except Exception as e:
            print(f"Error fetching data: {e}")
            # Return empty dataframe as fallback
            return pd.DataFrame(), {}
    
    def create_dashboard(self, df, metadata):
        """Create Excel dashboard with data"""
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # ===== DATA SHEET =====
        ws_data = wb.create_sheet("Budget Data")
        
        # Write data
        if not df.empty:
            # Write headers with formatting
            for col_num, column_name in enumerate(df.columns, 1):
                cell = ws_data.cell(row=1, column=col_num, value=column_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                ws_data.column_dimensions[get_column_letter(col_num)].width = 20
            
            # Write data rows
            for row_num, row_data in enumerate(df.itertuples(index=False), 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_data.cell(row=row_num, column=col_num, value=value)
            
            # Add auto-filter
            ws_data.auto_filter.ref = ws_data.dimensions
            
            # Add conditional formatting for status
            status_col = None
            for col_num, col_name in enumerate(df.columns, 1):
                if 'status' in col_name.lower():
                    status_col = get_column_letter(col_num)
                    break
            
            if status_col:
                # Format for different statuses
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                for row in range(2, len(df) + 2):
                    cell = ws_data[f"{status_col}{row}"]
                    if cell.value == "Approved":
                        cell.fill = green_fill
                    elif cell.value == "Pending":
                        cell.fill = yellow_fill
                    elif cell.value == "Rejected":
                        cell.fill = red_fill
        
        # ===== DASHBOARD SHEET =====
        ws_dash = wb.create_sheet("Dashboard")
        
        # Title
        ws_dash.merge_cells('A1:E1')
        title_cell = ws_dash['A1']
        title_cell.value = "CMM Budget Dashboard"
        title_cell.font = Font(size=20, bold=True, color="366092")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Last updated info
        ws_dash['A3'] = "Last Updated:"
        ws_dash['B3'] = metadata.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ws_dash['A4'] = "Total Records:"
        ws_dash['B4'] = metadata.get('total_records', 0)
        
        # Filter Controls
        ws_dash['A6'] = "FILTER CONTROLS"
        ws_dash['A6'].font = Font(bold=True, size=12)
        
        ws_dash['A7'] = "Region:"
        ws_dash['B7'] = self.filters['region']
        ws_dash['B7'].data_validation = {
            'type': 'list',
            'formula1': '"AMER,EMEA,APAC,ALL"',
            'allow_blank': True
        }
        
        ws_dash['A8'] = "Status:"
        ws_dash['B8'] = self.filters['status']
        ws_dash['B8'].data_validation = {
            'type': 'list',
            'formula1': '"Approved,Pending,Rejected,ALL"',
            'allow_blank': True
        }
        
        # Refresh Button Instructions
        ws_dash['A10'] = "TO REFRESH DATA:"
        ws_dash['A10'].font = Font(bold=True, color="FF0000")
        ws_dash['A11'] = "1. Update filters above"
        ws_dash['A12'] = "2. Go to Data tab → Refresh All"
        ws_dash['A13'] = "3. Or press Ctrl+Alt+F5"
        
        # Summary Statistics
        if not df.empty and 'amount_requested' in df.columns:
            ws_dash['A15'] = "FINANCIAL SUMMARY"
            ws_dash['A15'].font = Font(bold=True, size=12)
            
            stats = [
                ("Total Amount", df['amount_requested'].sum()),
                ("Average Amount", df['amount_requested'].mean()),
                ("Maximum Amount", df['amount_requested'].max()),
                ("Minimum Amount", df['amount_requested'].min()),
                ("Number of Vendors", df['vendor'].nunique() if 'vendor' in df.columns else 0)
            ]
            
            for i, (label, value) in enumerate(stats, 16):
                ws_dash[f'A{i}'] = label
                ws_dash[f'B{i}'] = value
                if 'Amount' in label:
                    ws_dash[f'B{i}'].number_format = '#,##0.00'
        
        # ===== CONFIGURATION SHEET =====
        ws_config = wb.create_sheet("Configuration")
        ws_config['A1'] = "API Configuration"
        ws_config['A2'] = "API URL:"
        ws_config['B2'] = self.api_url
        ws_config['A3'] = "Default Filters:"
        ws_config['B3'] = str(self.filters)
        ws_config['A4'] = "Created On:"
        ws_config['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Hide config sheet
        ws_config.sheet_state = 'hidden'
        
        # Set active sheet
        wb.active = ws_dash
        
        return wb
    
    def generate_template(self, output_path="Budget_Dashboard_Template.xlsx"):
        """Generate complete Excel template"""
        print("Fetching data from API...")
        df, metadata = self.fetch_data()
        
        print("Creating Excel dashboard...")
        wb = self.create_dashboard(df, metadata)
        
        print(f"Saving to {output_path}...")
        wb.save(output_path)
        
        # Create Power Query version
        self.create_power_query_template(df, "Budget_PowerQuery_Template.xlsx")
        
        print("✅ Template generation complete!")
        return output_path
    
    def create_power_query_template(self, df, output_path):
        """Create a version with embedded Power Query"""
        # This would require win32com or similar for full automation
        # For now, create a CSV for manual Power Query setup
        if not df.empty:
            df.to_csv("budget_data_sample.csv", index=False)
            print("📁 Sample CSV created for Power Query setup")
            
            # Create instruction file
            with open("PowerQuery_Setup.txt", "w") as f:
                f.write("""
POWER QUERY SETUP INSTRUCTIONS:
================================
1. Open Excel
2. Go to Data → Get Data → From File → From CSV
3. Select "budget_data_sample.csv"
4. In Power Query Editor:
   - Transform data as needed
   - Change data types
   - Add filters
5. Click Close & Load
6. To make it dynamic:
   - Right-click query → Advanced Editor
   - Replace file path with API URL:
     Source = Csv.Document(Web.Contents("https://eposaudioevents.com/budgets/excel_api.php?format=csv"))
================================
                """)

if __name__ == "__main__":
    generator = BudgetExcelGenerator()
    
    # Set custom filters if needed
    generator.filters = {
        'region': 'AMER',
        'status': 'ALL',
        'format': 'json'
    }
    
    generator.generate_template()